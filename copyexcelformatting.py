#!/usr/bin/env python
# coding: utf-8

# In[1]:


from copy import copy
import openpyxl
import os


#this program is used to copy the formatting of a column, row or entire sheet

def copy_cell_format(from_cell, to_cell):
    # this function copies a cell's format to antother cell
    
    if from_cell.has_style:
        to_cell.font = copy(from_cell.font)
        to_cell.border = copy(from_cell.border)
        to_cell.fill = copy(from_cell.fill)
        to_cell.number_format = copy(from_cell.number_format)
        to_cell.protection = copy(from_cell.protection)
        to_cell.alignment = copy(from_cell.alignment)
        
      
        
def copy_column_format(in_sheet, out_sheet = None, from_column_idx = None, to_column_idx = None):
    # this function copies the format from a column of a sheet to another column    
    
    if from_column_idx == None:
        # the default is to the use the last row of the table
        from_column_idx = in_sheet.max_column
        
    if to_column_idx == None:
        # the default is for the column format to be appended to the right of the table
        to_column_idx = from_column_idx + 1
        
    if out_sheet == None:
        # the default is for the out_sheet to be the same as the in_sheet
        
        out_sheet = in_sheet

    for each_row in list(range(1, in_sheet.max_row)):

        from_cell = in_sheet.cell(row = each_row, column = from_column_idx)
        to_cell = out_sheet.cell(row = each_row, column = to_column_idx)               
        
        if from_cell.has_style:
            copy_cell_format(from_cell, to_cell)
            
            
            
            
def copy_row_format(in_sheet, out_sheet = None, from_row_idx = None, to_row_idx = None):
    # this function copies the format from a row of a sheet to another row
    
    if from_row_idx == None:
        # the default is to the use the last row of the table
     
        from_row = in_sheet.max_row
        
    if to_row_idx == None:
        # the default is for the column format to be appended to the bottom of the table
        
        to_row_idx = from_row + 1
        
    if out_sheet == None:
        # the default is for the out_sheet to be the same as the in_sheet
        
        out_sheet = in_sheet
        
    for each_column in list(range(1, in_sheet.max_column)):
        
        from_cell = in_sheet.cell(row = from_row_idx, column = each_column)
        to_cell = out_sheet.cell(row = to_row_idx, column = each_column)               
        
        if from_cell.has_style:
            copy_cell_format(from_cell, to_cell)
            
            
            
def copy_formatting(from_sheet, to_sheet):
    
    # this may not work on styles that are defined in the workbook, but alas, it's something
    # another function should be created to examine all of the styles in the workbook and/or sheet
    
    row_idx = 1
    for row in from_sheet.iter_rows():
        copy_row_format(from_sheet, to_sheet, row_idx, row_idx)
        row_idx += 1
